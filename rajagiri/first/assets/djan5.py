import shutil
import MySQLdb
import boto3
import sys
import xlrd
import re
import os
import difflib
import pandas as pd
import openpyxl
import pathlib
import numpy
from openpyxl import Workbook
from sklearn.tree import DecisionTreeClassifier  # Import Decision Tree Classifier
# Import train_test_split function
from sklearn.model_selection import train_test_split
# Import scikit-learn metrics module for accuracy calculation
from sklearn import metrics
import pickle as pk
from difflib import get_close_matches
from difflib import SequenceMatcher

doc = sys.argv[1]

# doc = (doc.replace("/", "\\"))
doc = (doc.replace("['", ""))
doc = (doc.replace("']", ""))
doc = (doc.replace("'", ""))
doc = (doc.replace(" ", ""))

# print(doc)
doc = (doc.split(','))
mainpath = r"/usr/src/rajagiri"

otflag = 0  # FLAG variable for a particular type of report
i = "y"

#ot=input("Enter a name for output file: ")
# ot=ot+".txt"
#f1 = open(ot,"a")


# Establish a MySQL connection
database = MySQLdb.connect(
    host="localhost", user="root", passwd="mypass", db="rajagirihospital1")

cursor = database.cursor()

#selq = "SELECT distinct path from report where Patient_ID = %s"
selq1 = "select distinct documents from details where Patient_ID = %s"
pid1 = (sys.argv[2],)
cursor.execute(selq1, pid1)

myresult = list(cursor.fetchall())
print(myresult)
myresult = str(myresult)

myresult = myresult.replace('\\\\\\', "")
myresult = myresult.replace("((\" ", "")
myresult = myresult.replace("\",),)", "")

outs = re.findall(r'\'(.*?)\'', myresult)


#out = list(outs)
#outs = [item for t in myresult for item in t]
print("####", set(outs))
print("%%%", set(doc))


# print("####",set(myresult),"%%%%%")
# print(type(myresult))
# print("##",set(doc))
if(set(doc) == set(outs)):
    exit("Record Already Found")

selq2 = "DELETE from details where Patient_ID = %s"
pid2 = (sys.argv[2],)
cursor.execute(selq2, pid2)

cursor.close()

# Commit the transaction
database.commit()

# Close the database connection
database.close()


ot = sys.argv[2]+".txt"
f1 = open(ot, "a")

print(doc)
# while i=="y" or i=="Y":
for val in doc:
    #inval=input("Enter name of input file: ")
    #file = r"C:\python\Tesseract\Inputs\\"
    val = str(val)
    # val.replace('\\', '')
    # file = os.path.join(mainpath, val)
    file = mainpath+val
    with open(file, 'rb') as document:
        imageBytes = bytearray(document.read())
    #print(file+" loaded")
    # Amazon Textract client
    textract = boto3.client('textract')

    # Call Amazon Textract
    response = textract.detect_document_text(Document={'Bytes': imageBytes})

    # print(response)

    # Print detected text
    for item in response["Blocks"]:
        if item["BlockType"] == "LINE":
            f1.write(item["Text"] + '\n')
    f1.write("PAGE HAS ENDED" + '\n')
    #i=input("\nIs there a contination for the image(y/n): ")

f1.close()


################################ PART TWO EXXTRACTION  #######################

sdate = []
datel = []


###########extracting parameter and its type##################

wb = xlrd.open_workbook("/usr/src/rajagiri/first/assets/extract.xlsx")
sheet = wb.sheet_by_index(0)
sheet.cell_value(0, 0)
x = []
y = []
for i in range(sheet.nrows):
    x.append(sheet.cell_value(i, 1))
    y.append(sheet.cell_value(i, 2))
l = len(x)

for i in range(0, l):
    x[i] = x[i].upper()
    y[i] = y[i].upper()

#################################################################


f = open(ot, "r")            # ot contains textual content of image
eof = f.tell()
f.seek(0, 0)
t = f.readlines()
# f.close()
dt = ''.join(t)
##################date and time extraction########################
sdate = []
datel = []
stime = []
timel = []
dater = re.compile(
    "([0]?[1-9]|[1|2][0-9]|[3][0|1])[./-]([0]?[1-9]|[1][0-2])[./-]([0-9]{4}|[0-9]{2})")
sdate = re.findall(dater, dt)
if(sdate != []):
    for m in range(0, 3):
        datel.append(sdate[0][m])
        if(m != 2):
            datel.append("-")
    date = ''.join(datel)
    # print(date)  #contains date
timer = re.compile("([0-1]?\d|2[0-3])(?::([0-5]?\d))(?::([0-5]?\d))?")
stime = re.findall(timer, dt)
if(stime != []):
    for n in range(0, 3):
        if(stime[0][n] == ''):
            timel.append("00")
        else:
            timel.append(stime[0][n])
        if(n != 2):
            timel.append(":")
    time = ''.join(timel)
    # print(time)  #contains time
##########################################################


############removing \n ##########

t = [i.replace('\n', '') for i in t]
while("" in t):
    t.remove("")
tlen = len(t)

#################AGE AND SEX EXTRACTION#####################


age = 0  # Variable Stores Age of patient
sex = "Not specified"  # Variable stores sex of Patient
agse = ""
qqq = []    # temporary storage variable

for i in range(0, len(t)):  # finding the line of age and sex

    if(difflib.SequenceMatcher(None, str(t[i]).lower(), "age\sex").ratio() > 0.8):
        agse = str(t[i+1])
    if(difflib.SequenceMatcher(None, str(t[i]).lower(), "sex\age").ratio() > 0.8):
        agse = str(t[i+1])

p = re.compile('\d+')  # number extraction
if(p.findall(agse)):
    qqq = p.findall(agse)
    age = int(str(qqq[0]))


m = re.compile('.*male.*')  # male extraction

p = re.compile('.*female.*')  # male extraction
if(p.findall(agse.lower())):
    sex = "Female"

else:
    if(m.findall(agse.lower())):
        sex = "Male"


###############################################################
output = []
DOR = 0
SOR = 0
fintext = " "

###############paragragh type reports###############
if(output == []):

    #f = f.tell()
    # f.seek(0,0)
    # totem=f.readlines()
    # f.close()
    for i in range(0, len(t)):
        pika = []
        pika = re.split("\s+", t[i])
        if(pika[0].lower() == "diagnosis:" or pika[0].lower() == "diagnosis-" or pika[0].lower() == "diagnosis"):
            DOR = i

        if(len(pika) >= 3):
            if(pika[0].lower() == "when" and pika[1].lower() == "and" and pika[2].lower() == "how"):
                SOR = i+1

    fff = 0
    for i in range(DOR, SOR):
        ttt = 1
        if(re.match(".*\.com.*", str(t[i])) and fff == 0):
            fff = 1
        else:
            if(re.match(".*\.com.*", str(t[i])) and fff == 1):
                fff = 0
                ttt = 0

        if(fff == 0 and ttt == 1):
            fintext = fintext+"\n"+str(t[i])

if(fintext != " "):
    ot = "para"+ot
    # oteamf=open("paraoutput.txt","w")
    oteamf = open(ot, "w")
    oteamf.write(fintext)
    oteamf.close()
###############################################################
i = 0
f = 0
PPC = 0
PCT = []

##############extracting parameter from report#########

try:
    while(t[i] != eof):
        t[i] = t[i].upper()
        if(t[i] == "PAGE HAS ENDED"):
            PCT.append(PPC)
            PPC = 0
        for j in range(0, l):
            ne = 0
            seq = difflib.SequenceMatcher(None, t[i], x[j])
            d = seq.ratio()*100
            sd = []
            st = []
            mp = []
            f = 0
            v = 0
            timef = 0
            flag = 1
            if(d > 80):
                # print(t[i])
                ne = 1
                nd = []
                curr = i+1
                while(flag == 1):
                    out = []
                    f = 0
                    if(y[j] == "NUMERIC"):
                        sd = re.findall(dater, t[curr])
                        if(sd != []):
                            v = 1
                            st = re.findall(timer, t[curr])
                            if(st != []):
                                timef = 1
                        t[curr+v] = t[curr+v].replace(' A', '')
                        if(t[curr+v].replace('.', '').isdigit()):
                            if((curr+3+v) < tlen):
                                nd = re.findall(dater, t[curr+3+v])
                                mp = difflib.get_close_matches(t[curr+3+v], x)
                            f = 1
                            out.append(t[i])
                            out.append(t[curr+v])
                            if((len(t[curr+v+1]) < 5) or (t[curr+v+1].find('/') != -1)):
                                out.append(t[curr+v+1])
                    elif(y[j] == "LIST OF CHOICES"):
                        if(len(t[i+1].split()) == 1):
                            f = 1
                            out.append(t[i])
                            out.append(t[i+1])
                    else:
                        pass
                    if(f == 1):
                        if(v == 0):
                            out.append(date)
                            out.append(time)
                        else:
                            out.append(t[curr].split()[0])
                            if(timef == 1):
                                out.append(t[curr].split()[1])
                        output.append(out)
                        PPC = PPC+1
                        if(mp == []):
                            if(nd != []):
                                curr = curr+3+v
                            else:
                                flag = 0
                        else:
                            flag = 0
                    else:
                        flag = 0

            if(ne == 1):
                break
        i = i+1
except:
    pass


####################################################################

print("pct:", PCT)


if(output != []):

    #######################################################

    ########### ENTER INPUT TO EXCEL FILE RESULT ##############

    kkk = []
    df = pd.DataFrame()

    if(len(output[1]) == 5):  # UNIT SPECIFIED INPUT
        for i in range(0, len(output)):
            kkk.append(output[i][0])
            kkk.append(output[i][1])
            kkk.append(output[i][2])
            kkk.append(output[i][3])
            kkk.append(output[i][4])

    if(len(output[1]) == 4):  # UNIT NOT SPECIFIED
        for i in range(0, len(output)):
            kkk.append(output[i][0])
            kkk.append(output[i][1])
            kkk.append("Remarks")
            kkk.append(output[i][2])
            kkk.append(output[i][3])

    # Creating three columns
    df['Parameter'] = kkk[0::5]
    df['Value'] = kkk[1::5]
    df['Unit'] = kkk[2::5]
    df['Date'] = kkk[3::5]
    df['Time'] = kkk[4::5]
    df['Patient ID'] = sys.argv[2]  # Appending Hospital ID to the excel sheet

    df = df[['Patient ID', 'Parameter', 'Value', 'Unit', 'Date', 'Time']]
    # Converting to excel
    df.to_excel('/usr/src/rajagiri/first/assets/result.xlsx', index=False)

    ####################### PART THREE CLASSIFICATION  ################################

    wb = openpyxl.load_workbook('/usr/src/rajagiri/first/assets/lab1.xlsx')
    sheet = wb.active
    t = []
    for cellobj in list(sheet.columns)[1]:
        t.append(cellobj.value)
    t = t[1:]

    dict = {}

    for key, v1, v2 in sheet.iter_rows():
        if key.value in dict:
            dict[key.value].append(v1.value)
        else:
            dict[key.value] = [v1.value]

    parameters = []
    for cellobj in list(sheet.columns)[1]:
        parameters.append(cellobj.value)

    disparameters = list(set(parameters))

    features = disparameters

    flag = 0
    if flag:
        dict2 = {}

        for i in dict.keys():
            dict2[i] = []
            for j in disparameters:
                if j in dict[i]:
                    dict2[i].append(1)
                else:
                    dict2[i].append(0)

        disparameters.append("Target")

        df = pd.DataFrame(columns=disparameters)
        n = 0
        for i in dict2.keys():
            dict2[i].append(i)
            print(len(dict2[i]), " ", n)
            df.loc[n] = dict2[i]
            n += 1
        # print(df)
        file = open("/usr/src/rajagiri/first/assets/data_p", "wb")
        pk.dump(df, file)
    else:
        file = open("/usr/src/rajagiri/first/assets/data_p", "rb")
        df = pk.load(file)

    X = df[features]
    y = df.Target
    X_test = X
    y_test = y
    X_train = X
    y_train = y

    from sklearn.naive_bayes import GaussianNB

    gnb = GaussianNB()
    gnb.fit(X_train, y_train)

    # making predictions on the testing set
    # y_pred = gnb.predict(X_test)

    # comparing actual response values (y_test) with predicted response values (y_pred)
    from sklearn import metrics

    # print("Gaussian Naive Bayes model accuracy(in %):", metrics.accuracy_score(y_test, y_pred)*100)

    wb1 = openpyxl.load_workbook('/usr/src/rajagiri/first/assets/result.xlsx')
    sheet1 = wb1.active

    x = []
    list(sheet1.columns)[1]
    for cellobj in list(sheet1.columns)[1]:
        x.append(cellobj.value)
    x = x[1:]

    for i in range(0, len(x)):
        if 'RBC' in x[i]:
            x[i] = x[i].replace("RBC", "RBC COUNT")
        if 'BASO' in x[i]:
            x[i] = x[i].replace("BASO", "BASOPHILS")
        if 'MONO' in x[i]:
            x[i] = x[i].replace("MONO", "MONOCYTES")
        if 'LY' in x[i]:
            x[i] = x[i].replace("LY", "LYMPHOCYTES")
        if '#' in x[i]:
            x[i] = x[i].replace("#", "COUNT")
        if 'WBC TOTAL' in x[i]:
            x[i] = x[i].replace("WBC TOTAL", "NRBC COUNT / 100 WBC")
        if 'RDW - CV' in x[i]:
            x[i] = x[i].replace("RDW - CV", "RDW")
        if 'CREATININE' in x:
            if 'UREA' in x[i]:
                x[i] = x[i].replace("UREA", "BLOOD UREA")

        if 'TOTAL PROTEIN' in x[i]:
            x[i] = x[i].replace("TOTAL PROTEIN", "PROTEIN TOTAL")
        if 'SGOT' in x[i]:
            x[i] = x[i].replace("SGOT", "AST")
        if 'SGPT' in x[i]:
            x[i] = x[i].replace("SGPT", "ALT")
        if 'TOTAL BILIRUBIN' in x[i]:
            x[i] = x[i].replace("TOTAL BILIRUBIN", "BILIRUBIN TOTAL")
        if 'DIRECT BILIRUBIN' in x[i]:
            x[i] = x[i].replace("DIRECT BILIRUBIN", "BILIRUBIN DIRECT")
        if 'INDIRECT BILIRUBIN' in x[i]:
            x[i] = x[i].replace("INDIRECT BILIRUBIN", "BILIRUBIN INDIRECT")
    print(x)
    z = []
    for i in range(0, len(x)):
        z = (difflib.get_close_matches(x[i], t))
        try:
            print(x[i], "-->", z)
            x[i] = z[0]
        except:
            pass

    temp = x
    temp1 = x

    loop_count = 0

    while len(x) != 0:
        if (loop_count == 500):
            break
        test_x = []
        for j in features:
            if j in x:
                test_x.append(1)
            else:
                test_x.append(0)

        df1 = pd.DataFrame(columns=features)
        df1.loc[0] = test_x
        tx = df1[features]
        y_pred = gnb.predict(tx)
        abc = numpy.array_str(y_pred)
        abc = abc[2:len(abc) - 2]

        # print()
        for i in x:

            if (i in dict[abc]):
                print(i, "	-", abc)
                temp1[temp.index(i)] = abc

        x = [i for i in x if i not in dict[abc]]
        #x = list(set(x) - set(dict[abc]))
        if(abc in x):
            x.remove(abc)

        loop_count = loop_count + 1

    if (loop_count == 500):
        for i in range(0, len(x)):

            temp1[temp.index(x[i])] = "Test Name not Available"

    print(len(temp1))
    df = pd.read_excel('/usr/src/rajagiri/first/assets/result.xlsx')
    df['Test Name'] = temp1
    df.to_excel('/usr/src/rajagiri/first/assets/result.xlsx', index=False)

    ############################# PART FOUR RANGE FINDING AND IMPROPER MARKING #######

    def similar(a, b):  # function returns percentage(0-1) of similarity between two strings
        return SequenceMatcher(None, a, b).ratio()

    # opening result excel sheet

    tp = xlrd.open_workbook("/usr/src/rajagiri/first/assets/result.xlsx")
    tpsheet = tp.sheet_by_index(0)

    nrtp = tpsheet.nrows
    nctp = tpsheet.ncols
    nctp = nctp-1
    tname = []  # stores all testnames in index
    pname = []  # stores all parameter names in index
    reval = []  # stores value of each in index

    ### CODE TO STORE T&P names in variable  ###

    for i in range(0, nrtp):
        tname.append(tpsheet.cell_value(i, nctp))
        pname.append(tpsheet.cell_value(i, 1))
        reval.append(tpsheet.cell_value(i, 2))

    # Give the location of range execel file
    lo = "/usr/src/rajagiri/first/assets/ReferenceRange.xlsx"  # reference excel location

    # To open excel sheet of range
    wb = xlrd.open_workbook(lo)
    sheet = wb.sheet_by_index(0)

    nro = sheet.nrows  # total number of rows in rangereference excel

    range1 = []  # stores range of each rows by same index as in result excel

    # 0 th row in result excel's range will be stored in range1[0] and so on

    range2 = []  # stores normal or abnormal value of each row of result excel

    remlis = []  # stores remarks of each index

    for j in range(1, nrtp):
        vx = "Not Found"
        norm = "Not Found"
        f = 0
        for i in range(1, nro):
            if(sheet.cell_value(i, 0).lower() == tname[j].lower()):
                # check if similarity % > 70%
                if(similar(sheet.cell_value(i, 1).lower(), pname[j].lower()) > 0.7):
                    if(int(sheet.cell_value(i, 4)) == 0 and int(sheet.cell_value(i, 5)) == 0):
                        if(str(sheet.cell_value(i, 6)) == "Not specified"):
                            v1 = str(sheet.cell_value(i, 2))
                            v2 = str(sheet.cell_value(i, 3))
                            # normal or abnormal check
                            if(float(reval[j]) >= float(v1) and float(reval[j]) <= float(v2)):
                                norm = "Normal"
                            else:
                                norm = "Abnormal"
                            vx = " "
                            vx = v1+"-"+v2
                            remlis.append(sheet.cell_value(i, 7))
                            f = 1
                            break
                        if(str(sheet.cell_value(i, 6)).lower() == sex.lower()):
                            v1 = str(sheet.cell_value(i, 2))
                            v2 = str(sheet.cell_value(i, 3))
                            # normal or abnormal check
                            if(float(reval[j]) >= float(v1) and float(reval[j]) <= float(v2)):
                                norm = "Normal"
                            else:
                                norm = "Abnormal"
                            vx = " "
                            vx = v1+"-"+v2
                            remlis.append(sheet.cell_value(i, 7))
                            f = 1
                            break
                    else:
                        if(int(sheet.cell_value(i, 4)) <= age and int(sheet.cell_value(i, 5)) >= age):
                            v1 = str(sheet.cell_value(i, 2))
                            v2 = str(sheet.cell_value(i, 3))
                            # normal or abnormal check
                            if(float(reval[j]) >= float(v1) and float(reval[j]) <= float(v2)):
                                norm = "Normal"
                            else:
                                norm = "Abnormal"
                            vx = " "
                            vx = v1+"-"+v2
                            remlis.append(sheet.cell_value(i, 7))
                            f = 1
                            break
                        else:
                            remlis.append("Age out of range")
                            f = 1

        if(f == 0):
            remlis.append(" ")
        range1.append(vx)
        range2.append(norm)

    df = pd.read_excel('/usr/src/rajagiri/first/assets/result.xlsx')
    df['Range'] = range1
    df['Condition'] = range2
    df['Remarks'] = remlis
    df.to_excel('/usr/src/rajagiri/first/assets/result.xlsx', index=False)

    os.remove(ot.replace('para', ''))

    recordcount = 0
    #totalpara = 0
    path = []
    for p in PCT:
        for i in range(0, p):
            path.append(doc[recordcount])
        recordcount = recordcount + 1
        #totalpara = totalpara + p

    print(len(path))

# if(output!=[]):
    df = pd.read_excel('/usr/src/rajagiri/first/assets/result.xlsx')
    df['Path'] = path
    df.to_excel('/usr/src/rajagiri/first/assets/result.xlsx', index=False)

totalpara = 0
for p in PCT:
    totalpara = totalpara + p
print("totalpara,", totalpara)

if(output == []):
    print("Output is Null")


# Open the workbook and define the worksheet
book = xlrd.open_workbook("/usr/src/rajagiri/first/assets/result.xlsx")
sheet = book.sheet_by_name("Sheet1")

# Establish a MySQL connection
database = MySQLdb.connect(
    host="localhost", user="root", passwd="mypass", db="rajagirihospital1")

cursor = database.cursor()

delq = "DELETE from report where Patient_ID = %s"
pid = (sys.argv[2],)
cursor.execute(delq, pid)

# Create the INSERT INTO sql query
query = """INSERT INTO report (Patient_ID, Test_Parameter, Value, Unit, Date, Time, Test_Name, Original_Range, Conditions, Remarks , Path ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

query1 = """INSERT INTO details (Patient_ID, age, sex, documents) VALUES (%s, %s, %s, %s)"""

for r in range(1, sheet.nrows):
    Patient_ID = sheet.cell(r, 0).value
    Test_Parameter = sheet.cell(r, 1).value

    Unit = sheet.cell(r, 3).value
    Date = sheet.cell(r, 4).value
    Time = sheet.cell(r, 5).value
    Test_Name = sheet.cell(r, 6).value
    Original_Range = sheet.cell(r, 7).value
    Conditions = sheet.cell(r, 8).value
    Values = sheet.cell(r, 2).value
    Remarks = sheet.cell(r, 9).value
    Path = sheet.cell(r, 10).value

    # Assign values from each row
    values = (Patient_ID, Test_Parameter, Values, Unit, Date, Time,
              Test_Name, Original_Range, Conditions, Remarks, Path)

    cursor.execute(query, values)

value1 = (Patient_ID, str(age), sex, str(doc))
cursor.execute(query1, value1)

if(output == []):
    query2 = """INSERT INTO details (Patient_ID, age, sex, documents) VALUES (%s, %s, %s, %s)"""
    value2 = (sys.argv[2], str(age), sex, str(doc))
    cursor.execute(query2, value2)

cursor.close()

# Commit the transaction
database.commit()

# Close the database connection
database.close()
print("Details Successfully added to Rajagiri Hospital Database , Report table ")

if(age != 0):
    print("Age:", age)

print("Gender:", sex)

'''
import webbrowser

a_website = "C:\\python\\Python38\\result.xlsx"

# Open url in a new window of the default browser, if possible
webbrowser.open_new(a_website)
'''

excelsheet = sys.argv[2] + "result.xlsx"
src = "/usr/src/rajagiri/first/assets/result.xlsx"
#src = src + excelsheet
dest = "/usr/src/rajagiri/first/assets/Group2"
shutil.copy(src, dest)

src1 = "/usr/src/rajagiri/first/assets/result.xlsx"
desc = "/usr/src/rajagiri/first/assets/"
desc = desc + excelsheet
try:
    os.rename(src1, desc)
except:
    os.remove(desc)
    os.rename(src1, desc)

    ################################### END OF PROGRAM ############################
