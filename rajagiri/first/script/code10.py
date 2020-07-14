####################### ARGUMENT FROM PYCHARM ###########################################

import sys

#print(sys.argv[2])
doc = sys.argv[1]

doc = (doc.replace("/","\\"))
doc = (doc.replace("['",""))
doc = (doc.replace("']",""))
doc = (doc.replace("'",""))
doc = (doc.replace(" ",""))

#print(doc)
doc = (doc.split(','))


###########################################################################################

################# BASE CODE ################


#import cv2
import pandas as pd
import difflib
import openpyxl
import pytesseract
from pytesseract import image_to_string
from PIL import Image
from openpyxl import Workbook
#from pathlib import PurePath


##### To fetch the 2nd column in lab1 excel sheet #####

wb = openpyxl.load_workbook('C:\\python\\Python38\\lab1.xlsx')
sheet = wb.active
para=[]
for cellobj in list(sheet.columns)[1]:
        para.append(cellobj.value)
para=para[1:]

q='y'
ind=[]


details=[]
d=[]
norm="<>"
det=[]

#print(text)

patientid=sys.argv[2]

for val in doc :
	val = str(val)
	minus=0
	#val=input("\nEnter name of scanned image:")
	file_path0 = "C:\\Users\\user\\djangoproject\\rajagiri"
	
	file_path0 += val
	#print(file_path0)


	text=pytesseract.image_to_string(Image.open(file_path0),lang="eng")


	arr=["ABC","BIOCHEMISTRY","FFDF","CLINICAL PATHOLOGY","PQR","XYZ"]
	length=len(arr)
	#print(length)
	index=[]
	
	

	for i in range(0,length):
                 index.append(text.find(arr[i]))	
	#print(index)
	index.sort()
	for z in range(0,len(index)):
		if(index[z]==-1):
			minus=minus+1

	###################### For continuation page and report starts with parameters #########################################
			
	if(minus==len(arr)):
		x=list(text.split("\n"))
		#print(x)
		
		#x=(list(text.split()))

		for i in range(0,len(x)):
			if 'RBC' in x[i]:
				x[i]=x[i].replace("RBC","RBC COUNT")
			if 'BASO' in x[i]:
				x[i]=x[i].replace("BASO","BASOPHILS")
			if 'MONO' in x[i]:
				x[i]=x[i].replace("MONO","MONOCYTES")
			if 'LY' in x[i]:
				x[i]=x[i].replace("LY","LYMPHOCYTES")
			if '#' in x[i]:
				x[i]= x[i].replace("#"," COUNT")
		#print(x)
		z=[] 
		for i in range(0,len(x)):
			z=(difflib.get_close_matches(x[i],para))
			#print(x[i],"-->",z)
			if(len(z)!=0):
				x[i]=z[0]
				
				x[i]=x[i].split()
				
				x[i].insert(0,"#")
				#x[i].insert(len(x[i])-1,"\n")
				x[i]=" ".join(x[i])
				

		
		
		#print(",,,,,,,,,,,,,",x)
		
		text=(" ".join(x))
		#print(text)
		for z in range(0,len(para)):
			indtemp=text.find(para[z])
			if(indtemp>1):
				ind.append(indtemp)
		ind.sort()
		#print(ind)
		text1=text[ind[0]:]
		
		text1=list(text1.split("#"))
		#print("@@@@",text1)
		for i in range(0,len(text1)-2):
			if(len(text1[i])<10):
				print(text1[i])
				text1[i:i+2]=[''.join(text1[i:i+2])]
		#print("testing",text1)
		
		loop_end=len(text1)
		cnt=0
		b=0
		while b!=loop_end:
			if(b%2!=0):
				text1.insert(b,'\n')
				loop_end=loop_end+1
			cnt=cnt+1
			b=b+1
		
		for i in range(0,len(text1)):
			if(i%2==0):
				text1[i]=text1[i].strip()
		#print("MOdified",text1,"!!!!!")
		text1=",".join(text1)
		text1=text1.replace(",","")
		text1=text1.replace("  ","\n")
		text1="\n"+text1
		#print("string",text1)
		details.append(text1)
		count=count+1
		#q=input("Is there another image as continuation for the report?(y=yes,n=no):")
		#print("!!!!!!!!")
		#print(details)
		#continue

	############################# Second page ends  ###################################################################
	
		
		
	#print(index)
	print("\n") 
	j=0
	count=0
	for i in range(0,(length-1)):
		if(index[i]!=-1):
			details.append((text[index[i]:index[i+1]]))
			#print("\n\n")
			j=j+1
			count=count+1

	details.append(text[index[i+1]:]) 	
	count=count+1
	#print("first page completed")
	#q=input("Is there another image as continuation for the report?(y=yes,n=no):")
	#print(details)

####################################################################################################################################################

############# Writing the contents into output.txt file ###########

f1=open("C:\\python\\Python38\\output.txt","w")
t=" ".join(details)
f1.write(t)
f1.close()

####################################################################


###### To remove the blank line in output.txt file ######

fh = open("C:\\python\\Python38\\output.txt", "r")
lines = fh.readlines()
fh.close()
keep = []
for line in lines:	
	if not line.isspace():
		keep.append(line)


fh = open("C:\\python\\Python38\\output.txt", "w")
fh.write("".join(keep))
fh.close()

###########################################################


########## To remove unnecessery line #################
with open("C:\\python\\Python38\\output.txt", "r") as f:
    lines = f.readlines()
with open("output.txt", "w") as f:
    for line in lines:
        if line.strip("\n") != "LHR SION. 12:56:08PM 0.38 e073 /nb 0.1 - 0.6 DISCHARGE SUMMARY":
            f.write(line)

############################################################


f=open("C:\\python\\Python38\\output.txt","r")
#f.seek(-1,2)     # go to the file end.
eof = f.tell()   # get the end of file location
f.seek(0,0)      # go back to file beginning



t=f.readlines()
i=1
val=0
k=0
#print(t[0]);
#while(t[i]!="x\0c"):
#while(f.tell() != eof):

#print("\n\n");

##################### Comparing first line and the corresponding line and grouping the parameter names and its results ######################
print("Hello")
try:
	while(t[i]!= eof):
		#print(t[i].split())
		#t[i]=t[i]+t[i+1]
		
		if(t[i][0].isalpha()):
			abc=t[i]
			#print("abc  ",abc)

			###########################  Hardcoding #################################
			while(t[i+1][0].isdigit()):
				pqr=abc
				pqr=pqr+t[i+1]
				
				pqr=pqr.replace("\n"," ")
				#print("oooooooooooooooo",pqr)
				d=(list(pqr.split(" ")))

				# should be removed
				if(d[0]=="EOSIN" and d[-1]==""):
					d.pop()
				#
				
				d.pop()
				
				for p in range(0,len(d)):
					if d[p][0].isdigit():
						d[0 : p] = [' '.join(d[0 : p])]
						break
				for p in range(0,len(d)):
					if d[p][0] in norm:

						d[p : ] = [' '.join(d[p : ])]
						break
				for p in range(0,len(d)):
					if d[p][0]=='-' : 

						d[p-1 :p+2 ] = [' '.join(d[p-1 :p+2 ])]
						break

				for p in range(0,len(d)):
					if d[p][len(d[p])-1]=='-' : 

						d[p :p+2] = [' '.join(d[p :p+2 ])]
						break

				#for p in range(0,len(d)):
				if d[len(d)-1]==' ' :
					d.pop()

				for p in range(0,len(d)):
					if d[p][0]=='x' and len(d[p])<=2 : 

						d[p :p+2] = [' '.join(d[p :p+2 ])]
						break

				d.insert(0, patientid) 
				
				#print(d)
				i=i+1
				
				############ Entering the contents into det list ###########

				for x  in d:		
					det.append(x)

				############################################################
	
		############ Hardcoding ends ################
				
		i=i+1
		
except:
	print("terminated\n")
	
f.close()

####################################################################################################################################
#print(det)
df = pd.DataFrame() 
#print(len(det))  
##################### Putting the values into excel sheet ###############################

df['Patient ID']= det[0::7]
df['Test Parameter'] = det[1::7] 
df['Date'] = det[2::7] 
df['Time'] = det[3::7] 
df['Value'] = det[4::7]
df['Unit'] = det[5::7]
df['Range'] = det[6::7]

  
# Converting to excel 
df.to_excel('C:\\python\\Python38\\result.xlsx', index = False)
print("Hello")
###########################################################################################


#print("hello")

################### To find the Normal/Abnormal column ####################################

wb = openpyxl.load_workbook('C:\\python\\Python38\\result.xlsx')
sheet = wb.active
value=[]
rang=[]

#sheet.columns[3]
#list(sheet.columns)[3]
for cellobj in list(sheet.columns)[4]:
        value.append(cellobj.value)
for cellobj in list(sheet.columns)[6]:
        rang.append(cellobj.value)


#print(value)
#print(rang)
flag=0
normal=[]

for i in range(1,len(value)):
	if(rang[i][0]=='<' and rang[i][1]=='='):
		number=float(rang[i][2:].strip())
		if(float(value[i]) <= number):
			normal.append('Normal')
		else:
			normal.append('Abnormal')

	elif(rang[i][0]=='<'):
		number=float(rang[i][1:].strip())
		if(float(value[i]) < number):
			normal.append('Normal')		
		else:
			normal.append('Abnormal')
		#print(number)

	elif(rang[i][0]=='>' and rang[i][1]=='='):
		number=float(rang[i][2:].strip())
		if(float(value[i]) >= number):
			normal.append('Normal')
		else:
			normal.append('Abnormal')

	elif(rang[i][0]=='>'):
		number=float(rang[i][1:].strip())
		if(float(value[i]) > number):
			normal.append('Normal')
		else:
			normal.append('Abnormal')

	elif(rang[i][0]=='-'):
		number=float(rang[i])
		if(float(value[i])==number):
			normal.append('Normal')
		else:
 			normal.append('Abnormal')

	elif('-' in rang[i]):
		flag=1
		number1,number2=rang[i].split('-')
		try:
			number1=float(number1)
			number2=float(number2)
			#print("Check ",type(number1),type(number2))
			if(float(value[i]) >= number1 and float(value[i]) <= number2):
				normal.append('Normal')
			else:
				normal.append('Abnormal')
			#print("-",number1,number2)
		except:
			normal.append('Improper Value')
	
	else:
		normal.append('Improper Value')
	#print(normal)

#print(normal)


df['Normal'] = normal
df.to_excel('C:\\python\\Python38\\result.xlsx', index = False)

#####################################################################################################################
print("Hello")
################## To calculate the Efficiency ########################################################

wb = openpyxl.load_workbook('C:\\python\\Python38\\result.xlsx')
sheet = wb.active
normalvalues=[]
#total_cols=len(df.axes[1])
#print(total_cols)
for cellobj in list(sheet.columns)[7]:
        normalvalues.append(cellobj.value)
count_impval=0

for i in range(1,len(normalvalues)):
	if(normalvalues[i]!='Improper Value'):
		count_impval=count_impval+1
no_of_rows=len(normalvalues)-1
efficiency=(count_impval/no_of_rows)*100
print("No of proper values: ",count_impval)
print("No of rows: ",no_of_rows)
print("Efficiency: "+"{:.2f}".format(efficiency));
print("Hello")
##########################################################################################################


####################### CLASSIFICATION ######################################


import numpy
from sklearn.tree import DecisionTreeClassifier # Import Decision Tree Classifier
from sklearn.model_selection import train_test_split # Import train_test_split function
from sklearn import metrics #Import scikit-learn metrics module for accuracy calculation
import pickle as pk

wb = openpyxl.load_workbook('C:\\python\\Python38\\lab1.xlsx')
sheet = wb.active
t=[]
for cellobj in list(sheet.columns)[1]:
        t.append(cellobj.value)
t=t[1:]

dict={}

for key, v1, v2 in sheet.iter_rows():
	if key.value in dict:
		dict[key.value].append(v1.value)
	else:
	    	dict[key.value] = [v1.value]
#for i in dict.keys():
#	print(i,":",dict[i])


parameters=[]
for cellobj in list(sheet.columns)[1]:
        parameters.append(cellobj.value)

disparameters=list(set(parameters))

features=disparameters

flag = 0
if flag:
	dict2 = {}

	for i in dict.keys():
		dict2[i] = []
		for j in disparameters:
			if j in dict[i]:
				dict2[i].append(1)
			else:
				dict2[i].append(0) 

	#print(dict2["24 hour Urine  Albumin"])
	#for i in dict2.keys():
	#	print(i,":",dict2[i])
	disparameters.append("Target")
	print(len(disparameters))
	df = pd.DataFrame(columns=disparameters)
	n = 0
	for i in dict2.keys():
		dict2[i].append(i)
		print(len(dict2[i])," " ,n)
		df.loc[n] = dict2[i]
		n += 1
	print(df)
	file  = open("data_p","wb")
	pk.dump(df,file)
else:
	file  = open("C:\\python\\Python38\\data_p","rb")
	df = pk.load(file)



X=df[features]
#print(X)
y=df.Target
#print("!!!!!!!!!!!!!",y,len(y))



#X_train, X_test, y_train, y_test = cross_validation.train_test_split(X, y, train_size=0.8, stratify=y) # 70% training and 30% test
#X_train, X_test, y_train, y_test = train_test_split(X, y, train_size=0.9, random_state=1) # 70% training and 30% test


X_test = X
y_test = y
X_train = X
y_train = y

from sklearn.naive_bayes import GaussianNB 
gnb = GaussianNB() 
gnb.fit(X_train, y_train) 
  
# making predictions on the testing set 
#y_pred = gnb.predict(X_test) 
  
# comparing actual response values (y_test) with predicted response values (y_pred) 
from sklearn import metrics 
#print("Gaussian Naive Bayes model accuracy(in %):", metrics.accuracy_score(y_test, y_pred)*100)



wb1 = openpyxl.load_workbook('C:\\python\\Python38\\result.xlsx')
sheet1 = wb1.active
	
x=[]
list(sheet1.columns)[1]
for cellobj in list(sheet1.columns)[1]:
        x.append(cellobj.value)
x=x[1:]


'''
v=['HEMOGLOBIN','HEMOGLOBIN','HEMOGLOBIN','HEMOGLOBIN', 'NEUTRO', 'NEUTRO#', 'LYMP ;', 'LY#', 'MONO','MONO#','EOSIN','EOSIN#','BASO','BASO#','RBC',
'MCV','MCH','MCHC','RDW-CV','PLATELET COUNT','MPV'
]
for i in range(0,len(v)):
	x.append(v[i])
'''

'''
x=[
'BASOPHILS', 
'BASOPHILS COUNT',
'EOSINOPHILS',
'EOSINOPHILS COUNT',
'HAEMOGLOBIN',
'LYMPHOCYTES',
'LYMPHOCYTES COUNT',
'MCH',
'MCHC',
'MCV',
'MONOCYTES',
'MONOCYTES COUNT',
'MPV',
'NEUTROPHILS',
'NEUTROPHILS COUNT',
'PLATELET',
'RBC COUNT',
'RDW',
]
'''
#x.append('EOSINOPHILS')

for i in range(0,len(x)):
	if 'RBC' in x[i]:
		x[i]=x[i].replace("RBC","RBC COUNT")
	if 'BASO' in x[i]:
		x[i]=x[i].replace("BASO","BASOPHILS")
	if 'MONO' in x[i]:
		x[i]=x[i].replace("MONO","MONOCYTES")
	if 'LY' in x[i]:
		x[i]=x[i].replace("LY","LYMPHOCYTES")
	if '#' in x[i]:
		x[i]= x[i].replace("#","COUNT")
	if 'WEC TOTAL' in x[i]:
		x[i]=x[i].replace("WEC TOTAL","NRBC COUNT / 100 WBC")
z=[] 
for i in range(0,len(x)):
	z=(difflib.get_close_matches(x[i],t))
	#print(x[i],"-->",z)
	x[i]=z[0]

#print(x)
temp=x
temp1=x
#print(";;;;;;;;;;;;",temp)

loop_count=0


while len(x)!=0:
	if(loop_count==50):
		break;
	test_x = []
	for j in features:
		if j in x:
			test_x.append(1)
		else:
			test_x.append(0) 

	df1 = pd.DataFrame(columns=features)
	#print(df1)
	df1.loc[0] = test_x
	tx = df1[features]
	#print(tx)
	#print(df1[Target])


	y_pred = gnb.predict(tx)
	abc=numpy.array_str(y_pred)
	abc=abc[2:len(abc)-2] 
	#print(dict[abc])
	print()
	for i in x :
		#print(i)
		if(i in dict[abc]):
			#print(i,"	-",abc)	
			temp1[temp.index(i)]=abc
			#print(temp.index(i))
	x = [i for i in x if i not in dict[abc]] 
	#print("\n!!!!!!!",x)
	loop_count=loop_count+1
	
if(loop_count==50):
	for i in range(0,len(x)):
		#print(x[i],"       - Test Name not Available")
		temp1[temp.index(x[i])]="Test Name not Available"
#print("temp1",temp1)

df = pd.read_excel('C:\\python\\Python38\\result.xlsx')
df['Test Name'] = temp1
df.to_excel('C:\\python\\Python38\\result.xlsx', index = False)


################## DATABASE CODE ##########################


import xlrd
import MySQLdb

# Open the workbook and define the worksheet
book = xlrd.open_workbook("C:\\python\\Python38\\result.xlsx")
sheet = book.sheet_by_name("Sheet1")


# Establish a MySQL connection
database = MySQLdb.connect (host="localhost", user = "root", passwd = "nirmalsql", db = "RajagiriHospital")

cursor = database.cursor()

# Create the INSERT INTO sql query
query = """INSERT INTO report (Patient_ID, Test_Parameter, Date, Time, Value, Unit, Original_Range, Type, Test_Name) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"""

for r in range(1, sheet.nrows):
		Patient_ID	= sheet.cell(r,0).value
		Test_Parameter	= sheet.cell(r,1).value
		Date		= sheet.cell(r,2).value
		Time		= sheet.cell(r,3).value
		Value		= sheet.cell(r,4).value
		Unit		= sheet.cell(r,5).value
		Original_Range	= sheet.cell(r,6).value
		Type		= sheet.cell(r,7).value
		Test_Name	= sheet.cell(r,8).value
		

		# Assign values from each row
		values = (Patient_ID, Test_Parameter, Date, Time, Value, Unit, Original_Range, Type, Test_Name)

		cursor.execute(query, values)


cursor.close()

# Commit the transaction
database.commit()


# Close the database connection
database.close()
print("Details Successfully added to Rajagiri Hospital Database , Report table ")

# -*- coding: utf-8 -*-
import webbrowser
 
a_website = "C:\\python\\Python38\\result.xlsx"
 
# Open url in a new window of the default browser, if possible
webbrowser.open_new(a_website)



######## To print the contents in report table #########
'''
database = MySQLdb.connect (host="localhost", user = "root", passwd = "nirmalsql", db = "RajagiriHospital")
cursor = database.cursor()


cursor.execute(""" SELECT * FROM report """)
data = cursor.fetchall ()

for row in data :
    print(row)
database.close()
'''